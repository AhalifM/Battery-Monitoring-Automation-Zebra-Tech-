import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.exceptions import IllegalCharacterError
import pandas as pd
from datetime import timedelta
import threading
 
input_folder = ""
output_folder = ""
 
def count_files_to_convert(folder):
    total_files = 0
    for root, dirs, files in os.walk(folder):
        for filename in files:
            if filename.endswith(".txt"):
                total_files += 1
    return total_files
 
def convert_text_to_excel(progress_bar, total_files):
    global input_folder, output_folder
 
    if not output_folder:
        output_folder = os.path.join(os.path.dirname(input_folder), os.path.basename(input_folder) + "_converted")
 
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
 
    progress_increment = 100 / total_files
 
    for root, dirs, files in os.walk(input_folder):
        for filename in files:
            if filename.endswith(".txt"):
                input_file = os.path.join(root, filename)
                relative_path = os.path.relpath(input_file, input_folder)
                output_file = os.path.join(output_folder, relative_path.replace(".txt", ".xlsx"))
 
                output_dir = os.path.dirname(output_file)
                if not os.path.exists(output_dir):
                    os.makedirs(output_dir)
 
                convert_file(input_file, output_file)
 
                progress_bar['value'] += progress_increment
                progress_bar.update_idletasks()
 
    print("Conversion and conditional formatting completed.")
    generate_summary(output_folder)  # Automatically call generate_summary after conversion
 
def convert_file(input_file, output_file):
    wb = Workbook()
    ws = wb.active
 
    file_name = os.path.splitext(os.path.basename(input_file))[0]
    sanitized_name = "".join(c if c.isalnum() or c in "_ -%:." else "_" for c in file_name)
 
    try:
        ws.title = sanitized_name[:31]
    except IllegalCharacterError:
        ws.title = "Sheet1"
 
    with open(input_file, 'r') as infile:
        for row_index, line in enumerate(infile, start=1):
            values = line.strip().split(',')
            for col_index, value in enumerate(values, start=1):
                cell_value = value.strip("'")
                try:
                    cell_value = float(cell_value)
                except ValueError:
                    cell_value = "".join(c if c.isalnum() or c in "_ -%:." else "_" for c in cell_value)
 
                ws.cell(row=row_index, column=col_index, value=cell_value)
 
    apply_conditional_formatting(ws)
 
    wb.save(output_file)
 
def apply_conditional_formatting(ws):
    prev_color = None
 
    for row, (time_cell, value_cell) in enumerate(zip(ws['B'], ws['G']), start=1):
        if value_cell.value is not None:
            try:
                value = float(value_cell.value)
                if time_cell.value is not None:
                    time_str = str(time_cell.value)
                    if ":" in time_str:
                        hour, minute, second = map(int, time_str.split(":"))
                        if 10 <= hour < 17:
                            if value < 0:
                                value_cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                                value_cell.font = Font(color='FFFFFF')
                            elif value > 0:
                                value_cell.fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
                            else:
                                continue  # Skip zero values
 
            except ValueError:
                pass
 
def process_excel(file_path, output_ws, serial_number, file_name):
    # Read Excel file and process data
    first_row = pd.read_excel(file_path, nrows=1)
    unique_id = first_row.iloc[0, first_row.columns.get_loc('Unique ID')]
    unique_id = unique_id.split()[0]  # Trim unique ID after the first space
   
    header_df = pd.read_excel(file_path, nrows=5)
    header_row_index = header_df[header_df.iloc[:, 0] == 'Date'].index[0]
    df = pd.read_excel(file_path, skiprows=header_row_index + 1)
    df.columns = df.columns.str.strip()
    df['Time'] = df['Time'].str.strip()
    df['Time'] = pd.to_datetime(df['Time'], format='%H:%M:%S', errors='coerce')
    df = df.dropna(subset=['Time'])
    start_time = pd.to_datetime('10:00:00').time()
    end_time = pd.to_datetime('17:00:00').time()
    if not df.empty:
        df_after_10 = df[df['Time'].dt.time >= start_time]
        if not df_after_10.empty:
            start_time = df_after_10['Time'].min().time()
        filtered_df = df[df['Time'].dt.time <= end_time]
        if not filtered_df.empty:
            max_time = filtered_df['Time'].max().time()
            end_time = max_time
        df = df[(df['Time'].dt.time >= start_time) & (df['Time'].dt.time <= end_time)]
 
    if not df.empty:
        first_date = df['Date'].iloc[0]
        start_time_df = df[df['Time'].dt.time == start_time]['%Charge'].values
        end_time_df = df[df['Time'].dt.time == end_time]['%Charge'].values
        battery_start_percentage = int(start_time_df[0].rstrip('%')) if start_time_df.size > 0 else 'NULL'
        battery_end_percentage = int(end_time_df[0].rstrip('%')) if end_time_df.size > 0 else 'NULL'
        log_total_usage_hour = (pd.to_datetime(df['Date'].iloc[0] + ' ' + str(end_time)) - pd.to_datetime(df['Date'].iloc[0] + ' ' + str(start_time))).total_seconds() / 3600 if battery_start_percentage != 'NULL' and battery_end_percentage != 'NULL' else 'NULL'
        log_total_battery_power_drain = battery_start_percentage - battery_end_percentage if battery_start_percentage != 'NULL' and battery_end_percentage != 'NULL' else 'NULL'
        average_drain_per_hour = log_total_battery_power_drain / log_total_usage_hour if log_total_battery_power_drain != 'NULL' and log_total_usage_hour != 'NULL' else 'NULL'
        data = [serial_number, file_name, unique_id, first_date, start_time, f"{battery_start_percentage}%", end_time, f"{battery_end_percentage}%", log_total_usage_hour, log_total_battery_power_drain, average_drain_per_hour]
       
        # Check for gaps greater than 2 hours
        gap_exists = False
        if df.shape[0] > 1:
            time_diff = df['Time'].diff()
            gap_exists = (time_diff > timedelta(hours=2)).any()
       
        remarks = []
        if start_time == 'NULL':
            remarks.append('- Log file is outside of Operating Hours')
        if log_total_usage_hour != 'NULL' and log_total_usage_hour < 3:
            remarks.append('- Operating Hour less than 3 Hours')
        file_size_kb = os.path.getsize(file_path) / 1024  # Convert bytes to KB
        if file_size_kb < 1500:
            remarks.append('- File Size less than 1.5mb')
 
        # Count positive and negative values in 'Charge Rate' column between 10:00:00 and 17:00:00, excluding zero values
        positive_count = 0
        negative_count = 0
        for index, row in df.iterrows():
            value = row['Charge Rate']
            if isinstance(value, str):
                try:
                    num_value = float(value.rstrip('%'))
                    if num_value > 0:
                        positive_count += 1
                    elif num_value < 0:
                        negative_count += 1
                except ValueError:
                    continue
            elif isinstance(value, (int, float)):
                if value > 0:
                    positive_count += 1
                elif value < 0:
                    negative_count += 1
 
        # Check for negative values in 'Charge Rate' column
        try:
            if not (df['Charge Rate'].apply(lambda x: float(x.rstrip('%')) if isinstance(x, str) else x) < 0).any():
                remarks.append('- There is no negative battery charge')
        except TypeError:
            pass
 
        remarks_str = '\n'.join(remarks)
        data.extend([positive_count, negative_count, remarks_str])
 
        # Highlight file name in red if gap exists
        if gap_exists:
            data[1] = f'=HYPERLINK("{file_path}", "{file_name}")'
            data[-1] = '- Gap greater than 2 hours'
            output_ws.append(data)
            for cell in output_ws.iter_rows(min_row=output_ws.max_row, max_row=output_ws.max_row, min_col=2, max_col=2):
                cell[0].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        else:
            output_ws.append(data)
    else:
        data = [serial_number, file_name, unique_id, 'NULL', 'NULL', 'NULL', 'NULL', 'NULL', 'NULL', 'NULL', 'NULL', 0,
        0, 'Log file is outside of Operating Hours']
        output_ws.append(data)
 
 
       
def generate_summary(output_folder_path):
    print('The file processing starts now!')
    def process_folder(folder_path, output_ws):
        serial_number = ''.join(filter(str.isdigit, os.path.basename(folder_path)))
        output_file_exists = True
 
        # Process Excel files in the folder
        for file_name in os.listdir(folder_path):
            file_path = os.path.join(folder_path, file_name)
            if os.path.isdir(file_path):
                # If it's a directory, recursively process its contents
                process_folder(file_path, output_ws)
            elif file_name.endswith('.xlsx'):
                # Process Excel files
                if output_file_exists:
                    # Check if the serial number already exists in the output file
                    for row in output_ws.iter_rows(min_row=2, max_row=output_ws.max_row, min_col=1, max_col=1, values_only=True):
                        if serial_number in row:
                            print(f"The folder {folder_path} is already processed.")
                            return
                    output_file_exists = False
 
                process_excel(file_path, output_ws, serial_number, file_name)
 
    # Set the input folder path to the output folder from the main code
    main_folder_path = output_folder_path
 
    # Set the output file path for the summary
    summary_file_path = os.path.join(os.path.dirname(main_folder_path), "summary.xlsx")
 
    # Create a workbook and worksheet object using openpyxl
    wb = Workbook()
    ws = wb.active
 
    # Create the titles
    titles = ["Serial Number", "Filename", "Unique ID", "Date", "Start time", "Battery Start Percentage",
              "End time", "Battery End Percentage", "Log Total Usage Hour",
              "Total Power Drain", "Average Drain Per Hour", "Positive Count", "Negative Count", "Remarks"]
    ws.append(titles)
 
    for cell in ws[1]:
        cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
 
    # Process the main folder
    process_folder(main_folder_path, ws)
 
    # Apply center alignment to all cells in the worksheet, including headers
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Center alignment and wrap text
            if cell.row == 1:  # Header row
                cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # Keep header fill color
 
    # Auto-adjust column widths based on content
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except TypeError:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Add some extra width
        ws.column_dimensions[column].width = adjusted_width
 
    # Adjust Filename column width to be smaller
    ws.column_dimensions['B'].width = 25  # Adjust width as needed
 
    # Save the workbook
    wb.save(summary_file_path)
 
    print("Summary generation complete!")
    print(f"Output exported to: {summary_file_path}")
 
def browse_input():
    global input_folder
    input_folder = filedialog.askdirectory(title="Select Input Folder")
    input_folder_label.config(text=input_folder)
 
def browse_output():
    global output_folder
    output_folder = filedialog.askdirectory(title="Select Output Folder")
    if output_folder:
        output_folder_label.config(text=output_folder)
    else:
        output_folder_label.config(text="")
 
def convert():
    if not input_folder:
        messagebox.showerror("Error", "Please select an input folder.")
        return
 
    total_files = count_files_to_convert(input_folder)
    if total_files == 0:
        messagebox.showinfo("Info", "No .txt files found in the selected folder.")
        return
 
    progress_bar.mode = 'determinate'
    progress_bar.start()
 
    # Run conversion process in a separate thread
    conversion_thread = threading.Thread(target=convert_text_to_excel, args=(progress_bar, total_files))
    conversion_thread.start()
 
    # Check the status of the conversion thread periodically
    check_thread_status(progress_bar, conversion_thread)
 
def check_thread_status(progress_bar, thread):
    if thread.is_alive():
        # If the thread is still running, update progress bar and check again after 100 ms
        root.after(100, check_thread_status, progress_bar, thread)
    else:
        # If the thread has completed, stop the progress bar and show completion message
        progress_bar.stop()
        progress_bar.grid_forget()
        messagebox.showinfo("Success", "Conversion and file processing completed successfully.")
 
root = tk.Tk()
root.title("Text to Excel Converter")
 
input_frame = tk.Frame(root)
input_frame.pack(pady=10)
input_label= tk.Label(input_frame, text="Select Input Folder:")
input_label.pack(side=tk.LEFT)
input_folder_label = tk.Label(input_frame, text="")
input_folder_label.pack(side=tk.LEFT)
input_browse_button = tk.Button(input_frame, text="Browse", command=browse_input)
input_browse_button.pack(side=tk.LEFT)
 
output_frame = tk.Frame(root)
output_frame.pack(pady=10)
output_label = tk.Label(output_frame, text="Select Output Folder (optional):")
output_label.pack(side=tk.LEFT)
output_folder_label = tk.Label(output_frame, text="")
output_folder_label.pack(side=tk.LEFT)
output_browse_button = tk.Button(output_frame, text="Browse", command=browse_output)
output_browse_button.pack(side=tk.LEFT)
 
convert_button = tk.Button(root, text="Convert Text to Excel", command=convert)
convert_button.pack(pady=10)
 
progress_bar = ttk.Progressbar(root, mode='determinate', maximum=100)
progress_bar.pack(pady=10)
 
root.mainloop()